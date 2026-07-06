import io
import logging
import threading
import time
import unicodedata
from datetime import datetime, timedelta
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..extensions import db
from ..models.company import Company
from .boond_service import BoondService, BOOND_STATE_MAP

logger = logging.getLogger(__name__)

# ── Excel style constants ─────────────────────────────────────────────────────
_C_DARK_BLUE  = "0F2D55"
_C_MED_BLUE   = "1A4A8A"
_C_LIGHT_BLUE = "D6E4F7"
_C_WHITE      = "FFFFFF"
_C_TOTAL_BG   = "0F2D55"
_C_GOOD       = "C6EFCE"
_C_GOOD_FG    = "276221"
_C_MED        = "FFEB9C"
_C_MED_FG     = "9C5700"
_C_BAD        = "FFC7CE"
_C_BAD_FG     = "9C0006"
_C_GREY       = "F2F2F2"
_C_BORDER     = "B8CCE4"

_FONT_NORMAL  = Font(name="Calibri", size=10)
_FONT_BOLD    = Font(name="Calibri", size=10, bold=True)
_FONT_HEADER  = Font(name="Calibri", bold=True, color=_C_WHITE, size=11)
_FONT_TITLE   = Font(name="Calibri", bold=True, color=_C_WHITE, size=14)
_FONT_TOTAL   = Font(name="Calibri", bold=True, color=_C_WHITE, size=10)
_FONT_SUB     = Font(name="Calibri", size=9, color="595959")

_FILL_HEADER  = PatternFill("solid", fgColor=_C_DARK_BLUE)
_FILL_ACCENT  = PatternFill("solid", fgColor=_C_MED_BLUE)
_FILL_TOTAL   = PatternFill("solid", fgColor=_C_DARK_BLUE)
_FILL_EVEN    = PatternFill("solid", fgColor=_C_GREY)
_FILL_ODD     = PatternFill("solid", fgColor=_C_WHITE)
_FILL_GOOD    = PatternFill("solid", fgColor=_C_GOOD)
_FILL_MED     = PatternFill("solid", fgColor=_C_MED)
_FILL_BAD     = PatternFill("solid", fgColor=_C_BAD)

_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_RIGHT        = Alignment(horizontal="right",  vertical="center")

_THIN         = Side(style="thin",   color=_C_BORDER)
_THICK        = Side(style="medium", color=_C_DARK_BLUE)
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_TOP   = Border(left=_THIN, right=_THIN, top=_THICK, bottom=_THIN)

# Boond action type codes
ACTION_TYPES_RDV  = {60, 61, 62}
ACTION_TYPE_APPEL = 2

_COL_WIDTHS = [32, 13, 11, 13, 14, 10, 12, 11, 12]
_HEADERS    = [
    "Société", "Prospects", "Clients", "Partenaires",
    "Total contacts", "Nb RDV", "Taux RDV", "Nb Appels", "Taux Appel"
]

TOP_N_COMPANIES = 50

# ── Base data cache (Tier 1) ──────────────────────────────────────────────────
_BASE_CACHE: Dict = {}
_BASE_CACHE_VERSION = 1
_BASE_CACHE_MAX     = 8
_BASE_CACHE_TTL     = 3600


def _base_key(start: str, end: str) -> str:
    return f"{start}|{end}|bv{_BASE_CACHE_VERSION}"


def _base_cache_get(start: str, end: str):
    entry = _BASE_CACHE.get(_base_key(start, end))
    if entry and (datetime.now() - entry["ts"]).total_seconds() < _BASE_CACHE_TTL:
        age = int((datetime.now() - entry["ts"]).total_seconds())
        logger.warning("[BASE CACHE] HIT %s → %s (age %ds)", start, end, age)
        return entry["data"]
    return None


def _base_cache_set(start: str, end: str, data: dict):
    if len(_BASE_CACHE) >= _BASE_CACHE_MAX:
        oldest = min(_BASE_CACHE, key=lambda k: _BASE_CACHE[k]["ts"])
        del _BASE_CACHE[oldest]
        logger.warning("[BASE CACHE] EVICT oldest entry to make room")
    _BASE_CACHE[_base_key(start, end)] = {"data": data, "ts": datetime.now()}
    logger.warning("[BASE CACHE] SET %s → %s (%d companies, %d contacts, %d actions)",
                   start, end, len(data["company_meta"]),
                   len(data["all_contacts"]), len(data["all_actions"]))


def extract_address(attrs: Dict) -> str:
    parts = []
    for field in ("address", "street", "zipCode", "postalCode", "town", "city", "country"):
        v = attrs.get(field) or ""
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
    return ", ".join(parts) if parts else "N/A"


def _display_name(raw: str) -> str:
    if not raw:
        return raw
    stripped = raw.strip()
    if stripped == stripped.lower():
        return stripped.upper() if len(stripped) <= 6 else stripped.title()
    return stripped


def strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


def normalize_company_name(name: str) -> str:
    if not name:
        return "SOCIETE INCONNUE"
    n = strip_accents(name).upper()
    if "BNP" in n or "PARIBAS" in n:                  return "BNP PARIBAS"
    if "SOCIETE GENERALE" in n or "SOC GEN" in n:    return "SOCIETE GENERALE"
    if "NATIXIS" in n:                                return "NATIXIS"
    if "BPCE" in n:                                   return "GROUPE BPCE"
    if "BANQUE POPULAIRE" in n:                       return "GROUPE BPCE"
    if "CAISSE D'EPARGNE" in n:                       return "GROUPE BPCE"
    if "SNCF" in n:                                   return "SNCF"
    if "BOLLORE" in n:                                return "GROUPE BOLLORE"
    if "SAINT GOBAIN" in n or "ST GOBAIN" in n:      return "SAINT GOBAIN"
    if "TRANSACTIS" in n:                             return "BNP PARIBAS"
    if "AMUNDI" in n:                                 return "CREDIT AGRICOLE"
    if "CACIB" in n or "CREDIT AGRICOLE" in n:       return "CREDIT AGRICOLE"
    if "LCL" in n:                                    return "CREDIT AGRICOLE"
    if "AG2R" in n:                                   return "AG2R LA MONDIALE"
    if "EDF" in n:                                    return "EDF"
    if "ENGIE" in n or "GDF" in n:                   return "ENGIE"
    if "GRDF" in n:                                   return "GRDF"
    if "COVEA" in n:                                  return "COVEA"
    if "MALAKOFF" in n:                               return "MALAKOFF HUMANIS"
    if "CAISSE DES DEPOTS" in n or " CDC " in n:    return "CAISSE DES DEPOTS"
    if "AXA" in n:                                    return "AXA"
    if "CREDIT MUTUEL" in n or "ARKEA" in n:         return "CREDIT MUTUEL ARKEA"
    if "SACEM" in n:                                  return "SACEM"
    if "SGCIB" in n or "SG " in n:                   return "SOCIETE GENERALE"
    if "ALD AUTOMOTIVE" in n:                         return "SOCIETE GENERALE"
    return n.strip()


def _pct(nb: int, total: int) -> str:
    """Return 'XX%' capped at 100%, or 'N/A'."""
    if total <= 0:
        return "N/A"
    return f"{min(round(nb / total * 100), 100)}%"


class CompanyService:

    @staticmethod
    def sync_companies(
        start_date:    str,
        end_date:      str,
        contact_types: List[str] = None,
        company_names: List[str] = None,
        keyword:       str = "",
    ) -> List[Dict]:
        """
        KPI definitions:
          nb_contacts = contacts CREATED in [start_date, end_date] for this company
          nb_rdv      = unique contacts (any) that had ≥1 RDV action in the period
          nb_appel    = unique contacts (any) that had ≥1 appel action in the period

          taux_rdv   = min(nb_rdv,   nb_contacts) / nb_contacts  → shown as "67%"
          taux_appel = min(nb_appel, nb_contacts) / nb_contacts  → shown as "45%"
        """
        if contact_types is None:
            contact_types = ["Prospect", "Client", "Partenaire"]

        logger.warning("[SYNC] date=%s→%s | types=%s | companies=%s | kw=%r",
                       start_date, end_date, contact_types,
                       company_names or "all", keyword)

        # ── Tier 1: try base data cache ────────────────────────────────────────
        base = _base_cache_get(start_date, end_date)
        is_base_hit = base is not None

        if is_base_hit:
            company_meta       = base["company_meta"]
            all_contacts       = base["all_contacts"]
            contact_to_company = base["contact_to_company"]
            all_actions        = base["all_actions"]
            all_db_companies   = base["all_db_companies"]
        else:
            # ── BASE CACHE MISS: fetch everything from Boond + resolve names ──
            boond = BoondService()
            from .qwen_service import normalize_name, GROUP_RULES, get_canonical_names_batch

            all_db_companies = {str(c.boond_id): c for c in Company.query.all()}
            existing_groups  = list({c.group_name for c in all_db_companies.values() if c.group_name})

            raw_companies = boond.get_companies()
            company_meta: Dict[str, Dict] = {}
            unknowns: Dict[str, str] = {}

            for c in raw_companies:
                cid             = str(c.get("id"))
                attrs           = c.get("attributes", {})
                resolved_status = BOOND_STATE_MAP.get(attrs.get("state", 0), "Autre")
                raw_name        = attrs.get("name", f"id={cid}")

                name_lower = normalize_name(raw_name)
                canonical  = None
                for kw_rule, group in GROUP_RULES.items():
                    if kw_rule in name_lower:
                        canonical = group
                        break

                if canonical is None:
                    db_obj = all_db_companies.get(cid)
                    if db_obj and db_obj.group_name:
                        canonical = db_obj.group_name

                if canonical is None:
                    unknowns[cid] = raw_name
                    canonical = raw_name

                if canonical not in existing_groups:
                    existing_groups.append(canonical)

                company_meta[cid] = {
                    "canonical": canonical,
                    "status":    resolved_status,
                    "sector":    attrs.get("expertiseArea") or attrs.get("activitySector") or "N/A",
                    "address":   extract_address(attrs),
                    "raw_name":  raw_name,
                    "boond_id":  cid,
                }

            if unknowns:
                logger.warning("[SYNC] %d companies need Qwen AI (%d batch calls)",
                               len(unknowns), (len(unknowns) + 14) // 15)
                batch_result = get_canonical_names_batch(list(unknowns.values()), existing_groups)
                for cid, raw_name in unknowns.items():
                    canonical = batch_result.get(raw_name, raw_name)
                    company_meta[cid]["canonical"] = canonical
                    if canonical not in existing_groups:
                        existing_groups.append(canonical)
                    logger.warning("[QWEN][CONV] ✅ '%s' → '%s'", raw_name[:35], canonical[:35])

            # Fetch contacts (no date filter) — builds contact_id → company_id map
            all_contacts = boond._paginate("contacts", params={}, max_pages=5)
            contact_to_company: Dict[str, str] = {
                str(contact.get("id")): str(
                    contact["relationships"]["company"]["data"]["id"]
                )
                for contact in all_contacts
                if contact.get("relationships", {}).get("company", {}).get("data", {}).get("id")
            }

            # Fetch actions for the period
            all_actions = boond.get_all_actions(start_date=start_date, end_date=end_date)

            # Persist canonical names to DB
            for cid, meta in company_meta.items():
                db_obj = all_db_companies.get(cid)
                if db_obj:
                    db_obj.group_name = meta["canonical"]
                else:
                    db_obj = Company(
                        boond_id=cid, name=meta["raw_name"],
                        group_name=meta["canonical"], status=meta["status"],
                        sector=meta["sector"], address=meta["address"],
                    )
                    db.session.add(db_obj)
                    all_db_companies[cid] = db_obj
            db.session.commit()

            _base_cache_set(start_date, end_date, {
                "company_meta":       company_meta,
                "all_contacts":       all_contacts,
                "contact_to_company": contact_to_company,
                "all_actions":        all_actions,
                "all_db_companies":   all_db_companies,
            })

        # ── Step 3: count contacts created in date window ─────────────────────
        contacts_by_status: Dict[str, Dict[str, int]] = {}
        for contact in all_contacts:
            created_at = contact.get("attributes", {}).get("creationDate", "")
            if not (start_date <= created_at <= end_date):
                continue
            rel = contact.get("relationships", {}).get("company", {}).get("data")
            if not (rel and rel.get("id")):
                continue
            cid  = str(rel["id"])
            meta = company_meta.get(cid)
            if not meta:
                continue
            if meta["status"] not in contact_types:
                continue
            bucket = contacts_by_status.setdefault(cid, {})
            bucket[meta["status"]] = bucket.get(meta["status"], 0) + 1

        # ── Step 4: keyword filter ─────────────────────────────────────────────
        filtered_actions = all_actions
        if keyword.strip():
            kw_terms = [t.lower() for t in keyword.strip().split() if t]

            def _action_matches_keyword(action: dict) -> bool:
                attrs = action.get("attributes", {}) or {}
                text  = " ".join(filter(None, [
                    attrs.get("comment",     "") or "",
                    attrs.get("description", "") or "",
                    attrs.get("title",       "") or "",
                ])).lower()
                return all(term in text for term in kw_terms)

            filtered_actions = [a for a in all_actions if _action_matches_keyword(a)]
            logger.warning("[SYNC] Keyword filter '%s' → %d/%d actions kept",
                           keyword, len(filtered_actions), len(all_actions))

        rdv_actions   = [a for a in filtered_actions
                         if a.get("attributes", {}).get("typeOf") in ACTION_TYPES_RDV]
        appel_actions = [a for a in filtered_actions
                         if a.get("attributes", {}).get("typeOf") == ACTION_TYPE_APPEL]

        # ── Step 5: count UNIQUE contacts per company (deduplication fix) ──────
        # FIX vs original: instead of counting +1 per action, we track a SET of
        # unique contact_ids per company.  The returned count is the number of
        # DISTINCT contacts that had at least one action of that type.
        # This guarantees nb_rdv / nb_appel never counts the same person twice.

        def _unique_contacts_per_company(action_list: list) -> Dict[str, int]:
            company_to_contacts: Dict[str, set] = {}
            for action in action_list:
                rels = action.get("relationships") or {}
                # Try resolving via contact relationship keys
                contacts_found: Dict[str, str] = {}  # contact_id → company_id
                for key in ("contact", "contacts", "mainContact", "dependsOn"):
                    val = rels.get(key, {})
                    if not isinstance(val, dict):
                        continue
                    d     = val.get("data")
                    clist = d if isinstance(d, list) else ([d] if d else [])
                    
                    for cd in clist:
                        if cd and cd.get("id"):
                            # If using dependsOn, make sure it's actually pointing to a contact
                            if key == "dependsOn" and cd.get("type") and cd.get("type") != "contact":
                                continue
                            contact_id = str(cd["id"])
                            cid = contact_to_company.get(contact_id)
                            if cid:
                                contacts_found[contact_id] = cid
                    if contacts_found:
                        break
                # Fallback: direct company relationship on action
                if not contacts_found:
                    cd = rels.get("company", {}).get("data")
                    if cd and cd.get("id"):
                        cid = str(cd["id"])
                        action_id = str(action.get("id", "unknown"))
                        contacts_found[f"action_{action_id}"] = cid
                # Record each unique contact for its company
                for contact_id, cid in contacts_found.items():
                    company_to_contacts.setdefault(cid, set()).add(contact_id)
            return {cid: len(s) for cid, s in company_to_contacts.items()}

        rdv_per_company   = _unique_contacts_per_company(rdv_actions)
        appel_per_company = _unique_contacts_per_company(appel_actions)

        logger.warning("[SYNC] Dedup — %d companies with ≥1 RDV contact, %d with ≥1 appel contact",
                       len(rdv_per_company), len(appel_per_company))

        # ── Step 6: aggregate by canonical name ───────────────────────────────
        groups: Dict[str, Dict] = {}
        for cid, meta in company_meta.items():
            if meta["status"] not in contact_types:
                continue

            canonical      = meta["canonical"]
            bucket         = contacts_by_status.get(cid, {})
            nb_prospects   = bucket.get("Prospect",   0)
            nb_clients     = bucket.get("Client",     0)
            nb_partenaires = bucket.get("Partenaire", 0)
            nb_contacts    = nb_prospects + nb_clients + nb_partenaires

            # nb_rdv: unique contacts who had ≥1 RDV this month
            # nb_appel: unique contacts who had ≥1 appel this month (RAW — used as denominator)
            # taux_rdv   = nb_rdv   / nb_contacts  (how many new contacts got a meeting)
            # taux_appel = nb_rdv   / nb_appel      (of people called, how many got a meeting)
            nb_rdv   = min(rdv_per_company.get(cid, 0), nb_contacts)  # cap for taux_rdv
            nb_appel = appel_per_company.get(cid, 0)                   # raw for taux_appel denominator

            if canonical not in groups:
                groups[canonical] = {
                    "sub_companies":       [],
                    "sub_company_details": [],
                    "sector":              meta["sector"],
                    "address":             meta["address"],
                    "nb_prospects":        0,
                    "nb_clients":          0,
                    "nb_partenaires":      0,
                    "nb_contacts":         0,
                    "nb_rdv":              0,
                    "nb_appel":            0,
                }
            g = groups[canonical]
            g["sub_companies"].append(_display_name(meta["raw_name"]))
            g["nb_prospects"]   += nb_prospects
            g["nb_clients"]     += nb_clients
            g["nb_partenaires"] += nb_partenaires
            g["nb_contacts"]    += nb_contacts
            g["nb_rdv"]         += nb_rdv
            g["nb_appel"]       += nb_appel

            g["sub_company_details"].append({
                "name":        _display_name(meta["raw_name"]),
                "status":      meta["status"],
                "nb_contacts": nb_contacts,
                "nb_rdv":      nb_rdv,
                "nb_appel":    nb_appel,
                "taux_rdv":    _pct(nb_rdv,              nb_contacts),
                "taux_appel":  _pct(min(nb_rdv, nb_appel), nb_appel),  # rdv / appel
            })

            if not is_base_hit:
                company_obj = all_db_companies.get(cid)
                if company_obj:
                    company_obj.name        = meta["raw_name"]
                    company_obj.group_name  = canonical
                    company_obj.status      = meta["status"]
                    company_obj.sector      = meta["sector"]
                    company_obj.address     = meta["address"]
                    company_obj.nb_contacts = nb_contacts
                    company_obj.nb_rdv      = nb_rdv
                    company_obj.nb_appel    = nb_appel

        if not is_base_hit:
            db.session.commit()

        if company_names:
            canonical_filter = {n.upper() for n in company_names}
            groups = {k: v for k, v in groups.items() if k.upper() in canonical_filter}
            logger.warning("[SYNC] Company filter → %d groups kept", len(groups))

        # ── Step 7: build final result list ───────────────────────────────────
        all_results = []
        for canonical, g in groups.items():
            nb_contacts = g["nb_contacts"]
            # nb_rdv already capped at nb_contacts (per sub-company); safe to use directly
            nb_rdv      = g["nb_rdv"]
            # nb_appel is raw (not capped) — it is the denominator for taux_appel
            nb_appel    = g["nb_appel"]

            all_results.append({
                "name":                canonical,
                "sub_companies":       sorted(set(g["sub_companies"])),
                "sub_company_details": sorted(
                    g["sub_company_details"],
                    key=lambda x: x["nb_contacts"], reverse=True
                ),
                "address":             g["address"],
                "sector":              g["sector"],
                "nb_prospects":        g["nb_prospects"],
                "nb_clients":          g["nb_clients"],
                "nb_partenaires":      g["nb_partenaires"],
                "nb_contacts":         nb_contacts,
                "nb_rdv":              nb_rdv,
                "nb_appel":            nb_appel,
                "taux_rdv":            _pct(nb_rdv, nb_contacts),              # rdv / contacts
                "taux_appel":          _pct(min(nb_rdv, nb_appel), nb_appel),  # rdv / appel
            })

        all_results = [r for r in all_results
                       if r["nb_contacts"] > 0 or r["nb_rdv"] > 0 or r["nb_appel"] > 0]
        all_results.sort(key=lambda x: x["nb_contacts"], reverse=True)
        return all_results[:TOP_N_COMPANIES]


    # ── Contact stats with AI deduplication (colleague's work) ────────────────

    @staticmethod
    def get_contacts_stats() -> Dict:
        from .boond_service import get_companies_dict, get_all_contacts as _get_all_contacts
        from .qwen_service import get_canonical_name, groupe_par_nom, ia_cache

        companies_dict = get_companies_dict()

        today      = datetime.now()
        year, month = today.year, today.month

        if today.day <= 7:
            if month == 1:
                month = 12
                year -= 1
            else:
                month -= 1

        first_day  = datetime(year, month, 1)
        if month == 12:
            last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(year, month + 1, 1) - timedelta(days=1)

        contacts   = _get_all_contacts(first_day.strftime("%Y-%m-%d"), last_day.strftime("%Y-%m-%d"))
        month_name = first_day.strftime("%B %Y")

        all_db_companies  = {str(c.boond_id): c for c in Company.query.all()}
        groupes_existants = [c.group_name for c in all_db_companies.values() if c.group_name]
        stats      = {}
        variations = {}

        for contact in contacts:
            try:
                company_id = None
                if "relationships" in contact and "company" in contact["relationships"]:
                    company_data = contact["relationships"]["company"].get("data")
                    if company_data:
                        company_id = str(company_data.get("id"))
                if not company_id or company_id not in companies_dict:
                    continue

                company_info  = companies_dict[company_id]
                original_name = company_info["name"]
                company_state = company_info["state"]

                groupe = get_canonical_name(original_name, list(set(groupes_existants)))
                if groupe not in groupes_existants:
                    groupes_existants.append(groupe)

                variations.setdefault(groupe, set()).add(original_name)

                if groupe not in stats:
                    stats[groupe] = {
                        "display_name": groupe,
                        "prospects": 0,
                        "clients":   0,
                        "partners":  0,
                        "total":     0,
                    }

                state_lower = company_state.lower()
                if "prospect" in state_lower:
                    stats[groupe]["prospects"] += 1
                    stats[groupe]["total"]     += 1
                elif "client" in state_lower:
                    stats[groupe]["clients"] += 1
                    stats[groupe]["total"]   += 1
                elif "partenaire" in state_lower:
                    stats[groupe]["partners"] += 1
                    stats[groupe]["total"]    += 1
            except Exception:
                continue

        db_conversion: Dict[str, Dict] = {}
        for db_obj in all_db_companies.values():
            grp = db_obj.group_name or normalize_company_name(db_obj.name or "")
            if grp not in db_conversion:
                db_conversion[grp] = {"nb_rdv": 0, "nb_appel": 0, "nb_contacts": 0}
            db_conversion[grp]["nb_rdv"]      += db_obj.nb_rdv      or 0
            db_conversion[grp]["nb_appel"]    += db_obj.nb_appel    or 0
            db_conversion[grp]["nb_contacts"] += db_obj.nb_contacts or 0

        result = []
        for name, data in stats.items():
            if data["total"] == 0:
                continue
            conv = db_conversion.get(name, {})
            nb_rdv      = conv.get("nb_rdv", 0)
            nb_appel    = conv.get("nb_appel", 0)
            nb_contacts = conv.get("nb_contacts", 0)
            result.append({
                "display_name": name,
                "prospects":    data["prospects"],
                "clients":      data["clients"],
                "partners":     data["partners"],
                "total":        data["total"],
                "other_names":  list(variations.get(name, set()))[:5],
                "nb_rdv":       nb_rdv,
                "nb_appel":     nb_appel,
                "taux_rdv":     _pct(nb_rdv,   nb_contacts),
                "taux_appel":   _pct(nb_appel, nb_contacts),
            })

        result.sort(key=lambda x: x["total"], reverse=True)
        return {"month": month_name, "stats": result}

    # ── Excel export ──────────────────────────────────────────────────────────

    @staticmethod
    def build_xlsx_from_db(rows: list = None, report_rows: list = None) -> bytes:
        if rows is None:
            companies = (
                Company.query
                .filter(Company.nb_contacts > 0)
                .order_by(Company.nb_contacts.desc())
                .limit(TOP_N_COMPANIES)
                .all()
            )
            rows = []
            for c in companies:
                nb_contacts = c.nb_contacts or 0
                nb_rdv      = c.nb_rdv or 0
                nb_appel    = c.nb_appel or 0
                rows.append({
                    "name":           c.name,
                    "nb_prospects":   nb_contacts if c.status == "Prospect"   else 0,
                    "nb_clients":     nb_contacts if c.status == "Client"     else 0,
                    "nb_partenaires": nb_contacts if c.status == "Partenaire" else 0,
                    "nb_contacts":    nb_contacts,
                    "nb_rdv":         nb_rdv,
                    "taux_rdv":       _pct(nb_rdv,   nb_contacts),
                    "nb_appel":       nb_appel,
                    "taux_appel":     _pct(nb_appel, nb_contacts),
                })

        xlsx_rows = []
        for r in rows:
            nb_contacts    = r.get("nb_contacts", 0)
            nb_rdv         = r.get("nb_rdv", 0)
            nb_appel       = r.get("nb_appel", 0)
            taux_rdv_str   = str(r.get("taux_rdv",   "N/A"))
            taux_appel_str = str(r.get("taux_appel", "N/A"))

            xlsx_rows.append({
                "société":        r.get("name", ""),
                "prospects":      r.get("nb_prospects", 0),
                "clients":        r.get("nb_clients", 0),
                "partenaires":    r.get("nb_partenaires", 0),
                "total_contacts": nb_contacts,
                "nb_rdv":         nb_rdv,
                "taux_rdv":       taux_rdv_str,
                "nb_appel":       nb_appel,
                "taux_appel":     taux_appel_str,
            })

        return CompanyService._make_xlsx(xlsx_rows, report_rows=report_rows)

    @staticmethod
    def _make_xlsx(rows: list, report_rows: list = None) -> bytes:
        import io as _io
        from datetime import datetime as _dt
        from openpyxl.utils import get_column_letter as _gcl

        def _rate_tier(val_str: str):
            # val_str is like "67%" or "N/A"
            try:
                v = float(str(val_str).replace("%", "").strip())
            except (ValueError, TypeError):
                return None, None
            if v == 0:   return _FILL_BAD,  _C_BAD_FG
            if v < 20:   return _FILL_BAD,  _C_BAD_FG
            if v < 50:   return _FILL_MED,  _C_MED_FG
            return _FILL_GOOD, _C_GOOD_FG

        totals = {
            "société":        "TOTAL",
            "prospects":      sum(r["prospects"]      for r in rows),
            "clients":        sum(r["clients"]        for r in rows),
            "partenaires":    sum(r["partenaires"]    for r in rows),
            "total_contacts": sum(r["total_contacts"] for r in rows),
            "nb_rdv":         sum(r["nb_rdv"]         for r in rows),
            "taux_rdv":       "",
            "nb_appel":       sum(r["nb_appel"]       for r in rows),
            "taux_appel":     "",
        }

        wb  = Workbook()
        ws  = wb.active
        ws.title = "Tableau de bord"
        ws.sheet_view.showGridLines = False

        now_str = _dt.now().strftime("%d/%m/%Y à %Hh%M")

        for col in range(1, len(_HEADERS) + 1):
            ws.column_dimensions[_gcl(col)].width = _COL_WIDTHS[min(col - 1, len(_COL_WIDTHS) - 1)]

        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 14

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_HEADERS))
        title_cell = ws.cell(row=1, column=1, value="📊  Pilotis — Tableau de conversion commerciale")
        title_cell.font = _FONT_TITLE; title_cell.fill = _FILL_HEADER; title_cell.alignment = _LEFT

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(_HEADERS))
        sub_cell = ws.cell(row=2, column=1, value=f"Export généré le {now_str}")
        sub_cell.fill = _FILL_ACCENT; sub_cell.alignment = _LEFT
        sub_cell.font = Font(name="Calibri", size=9, color=_C_LIGHT_BLUE)

        ws.row_dimensions[3].height = 6

        HDR_ROW = 4
        for col_idx, h in enumerate(_HEADERS, start=1):
            cell = ws.cell(row=HDR_ROW, column=col_idx, value=h)
            cell.font = _FONT_HEADER; cell.fill = _FILL_HEADER
            cell.alignment = _CENTER; cell.border = _BORDER
        ws.row_dimensions[HDR_ROW].height = 28

        all_rows      = rows + [totals]
        total_row_idx = HDR_ROW + len(all_rows)

        for i, row in enumerate(all_rows, start=HDR_ROW + 1):
            is_total = (i == total_row_idx)
            is_even  = ((i - HDR_ROW) % 2 == 0)

            ws.append([
                row["société"], row["prospects"], row["clients"], row["partenaires"],
                row["total_contacts"], row.get("nb_rdv", ""), row.get("taux_rdv", ""),
                row.get("nb_appel", ""), row.get("taux_appel", ""),
            ])
            ws.row_dimensions[i].height = 20

            rdv_fill,   rdv_fg   = _rate_tier(row.get("taux_rdv",   ""))
            appel_fill, appel_fg = _rate_tier(row.get("taux_appel", ""))

            for col_idx in range(1, len(_HEADERS) + 1):
                cell = ws.cell(row=i, column=col_idx)
                if is_total:
                    cell.font = _FONT_TOTAL; cell.fill = _FILL_TOTAL
                    cell.alignment = _CENTER if col_idx > 1 else _LEFT
                    cell.border = Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THICK)
                else:
                    cell.font = _FONT_BOLD if col_idx == 1 else _FONT_NORMAL
                    cell.fill = _FILL_EVEN if is_even else _FILL_ODD
                    cell.alignment = _LEFT if col_idx == 1 else _CENTER
                    cell.border = _BORDER
                    if col_idx == 7 and rdv_fill:
                        cell.fill = rdv_fill
                        cell.font = Font(name="Calibri", size=10, bold=True, color=rdv_fg)
                    if col_idx == 9 and appel_fill:
                        cell.fill = appel_fill
                        cell.font = Font(name="Calibri", size=10, bold=True, color=appel_fg)
                    if col_idx == 2 and row["prospects"]   > 0:
                        cell.font = Font(name="Calibri", size=10, bold=True, color="1A4A8A")
                    if col_idx == 3 and row["clients"]     > 0:
                        cell.font = Font(name="Calibri", size=10, bold=True, color="1D6E35")
                    if col_idx == 4 and row["partenaires"] > 0:
                        cell.font = Font(name="Calibri", size=10, bold=True, color="6B3FA0")

        ws.freeze_panes = f"A{HDR_ROW + 1}"
        ws.auto_filter.ref = f"A{HDR_ROW}:{_gcl(len(_HEADERS))}{HDR_ROW}"

        legend_row = total_row_idx + 2
        ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=len(_HEADERS))
        leg = ws.cell(row=legend_row, column=1,
                      value="Légende   🟢 Bon (≥50%)   🟠 Moyen (20%-50%)   🔴 Faible (<20%)   ⚪ Nul (0%)")
        leg.font = Font(name="Calibri", size=9, italic=True, color="595959")
        leg.alignment = _LEFT

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()